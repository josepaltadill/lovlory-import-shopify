from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


BASE = Path(__file__).resolve().parents[1]
SOURCE = BASE / "LovLory_estructura_colecciones_shopify_v4.xlsx"
OUTPUT = BASE / "LovLory_estructura_colecciones_shopify_MVP_v1.xlsx"

HEADER_FIELDS = [
    "Subtítulo superior cabecera",
    "Descripción corta cabecera",
]


def replace_text(value: object) -> object:
    if isinstance(value, str):
        return (
            value.replace("v3", "MVP v1")
            .replace("v2", "v4")
            .replace("Decisión aplicada en v3", "Decisión aplicada en MVP v1")
        )
    return value


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def delete_unwanted_rows(ws, key_col: int, keep_values: set[str], header_row: int = 3) -> None:
    for row in range(ws.max_row, header_row, -1):
        value = ws.cell(row, key_col).value
        if value not in keep_values:
            ws.delete_rows(row, 1)


def find_row(ws, key_col: int, value: str, header_row: int = 3) -> int | None:
    for row in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row, key_col).value == value:
            return row
    return None


def ensure_brand(ws, brand: str, handle: str, categories: str, filters: str, notes: str) -> None:
    if find_row(ws, 1, brand):
        return
    template_row = min(ws.max_row, 4)
    target = ws.max_row + 1
    copy_row_style(ws, template_row, target)
    values = [
        brand,
        handle,
        "Sí",
        f"vendor = {brand}",
        categories,
        filters,
        notes,
    ]
    for col, value in enumerate(values, 1):
        ws.cell(target, col).value = value


def ensure_header_fields(ws) -> None:
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
    missing_headers = [field for field in HEADER_FIELDS if field not in headers]
    if not missing_headers:
        return

    template_col = ws.max_column
    template_data_row = min(ws.max_row, 4)

    for field in missing_headers:
        new_col = ws.max_column + 1
        ws.cell(3, new_col).value = field

        for row in range(1, ws.max_row + 1):
            src = ws.cell(row, template_col)
            dst = ws.cell(row, new_col)
            if src.has_style:
                dst._style = copy(src._style)
            if src.number_format:
                dst.number_format = src.number_format
            if src.alignment:
                dst.alignment = copy(src.alignment)
            if src.font:
                dst.font = copy(src.font)
            if src.fill:
                dst.fill = copy(src.fill)
            if src.border:
                dst.border = copy(src.border)

        ws.cell(3, new_col).value = field
        ws.cell(template_data_row, new_col).value = ""
        column_letter = ws.cell(3, new_col).column_letter
        ws.column_dimensions[column_letter].width = max(
            ws.column_dimensions[ws.cell(3, template_col).column_letter].width or 18,
            24,
        )


def write_decision_sheet(wb) -> None:
    if "Decisiones MVP" in wb.sheetnames:
        del wb["Decisiones MVP"]
    ws = wb.create_sheet("Decisiones MVP", 0)
    ws.sheet_view.showGridLines = False

    rows = [
        ["LovLory - Especificación Shopify MVP v1", "", "", ""],
        ["Objetivo", "Crear ahora las colecciones principales, solo Novedades/Ofertas como secundarias, y marcas ORGIE, SVAKOM, TENGA y MISTRESS.", "", ""],
        ["", "", "", ""],
        ["Área", "Decisión MVP", "Cómo implementarlo en Shopify", "Notas operativas"],
        [
            "Colecciones principales",
            "Crear las colecciones principales del v4.",
            "Preferir colecciones automáticas cuando la regla dependa de tag, product_type, vendor o metafield; usar manual si la selección editorial no se puede inferir.",
            "No crear todavía microcolecciones no validadas por stock real.",
        ],
        [
            "Secundarias - Novedades",
            "Crear ahora.",
            "Recomendado: colección automática por tag = novedad, ordenada por CREATED_DESC. Alternativa: colección manual editorial.",
            "Shopify no siempre permite una regla nativa tipo 'últimos N productos' como WordPress. Lo más controlable es etiquetar productos nuevos durante importación/publicación y retirar el tag cuando toque.",
        ],
        [
            "Secundarias - Ofertas",
            "Crear ahora.",
            "Recomendado: colección automática por descuento si la condición de compare_at_price está disponible; fallback robusto: tag = oferta.",
            "Además puede existir filtro de precio/descuento en la colección, pero una colección Ofertas propia ayuda a SEO, home, menú y campañas.",
        ],
        [
            "Resto secundarias",
            "No crear por ahora.",
            "Mantener como ideas para campañas, landings o filtros futuros.",
            "Evita sobrecargar menú y trabajo de clasificación inicial.",
        ],
        [
            "Marcas",
            "Crear solo ORGIE, SVAKOM, TENGA y MISTRESS.",
            "Colecciones automáticas por vendor exacto.",
            "Normalizar mayúsculas/minúsculas durante importación para que las reglas funcionen.",
        ],
        [
            "Tags, filtros y metafields",
            "Definir desde cero una taxonomía limpia.",
            "No confiar en el etiquetado actual de WordPress si viene de una importación antigua.",
            "Primero importar/exportar datos, limpiar columnas maestras y luego alimentar tags/metafields de Shopify.",
        ],
        [
            "Exportación de productos",
            "Si el exportador no filtra por marca, exportar todos.",
            "Después filtramos/limpiamos por marca, categoría, atributos o meta campos en Excel/CSV.",
            "Exportar todo suele ser más seguro que sacar subconjuntos incompletos por categoría.",
        ],
    ]
    for r, row in enumerate(rows, 1):
        for c, value in enumerate(row, 1):
            ws.cell(r, c).value = value

    title_fill = PatternFill("solid", fgColor="4A102A")
    header_fill = PatternFill("solid", fgColor="6B2143")
    note_fill = PatternFill("solid", fgColor="F6EEF2")
    white = "FFFFFF"
    thin = Side(style="thin", color="E6CED8")

    ws.merge_cells("A1:D1")
    ws["A1"].font = Font(bold=True, color=white, size=15)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("B2:D2")
    ws["A2"].font = Font(bold=True)
    ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["B2"].fill = note_fill
    ws.row_dimensions[2].height = 42

    for cell in ws[4]:
        cell.font = Font(bold=True, color=white)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {"A": 26, "B": 34, "C": 52, "D": 52}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 58
    ws.freeze_panes = "A5"


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = replace_text(cell.value)

    resumen = wb["Resumen"]
    resumen["A1"] = "LovLory - Arquitectura Shopify MVP v1 desde v4"
    resumen["B4"] = "MVP v1"
    resumen["C4"] = "Versión operativa para crear colecciones principales, Novedades/Ofertas y marcas priorizadas."

    secundarias = wb["Colecciones secundarias"]
    delete_unwanted_rows(secundarias, 1, {"Novedades", "Ofertas"})
    novedades_row = find_row(secundarias, 1, "Novedades")
    if novedades_row:
        secundarias.cell(novedades_row, 4).value = "Sí"
        secundarias.cell(novedades_row, 6).value = "tag = novedad; ordenar por fecha de creación descendente"
        secundarias.cell(novedades_row, 10).value = (
            "Crear como colección viva. Shopify no siempre replica 'últimos añadidos' por fecha como regla automática; "
            "lo más fiable es taggear novedades en importación/publicación y retirar el tag cuando deje de ser novedad."
        )
    ofertas_row = find_row(secundarias, 1, "Ofertas")
    if ofertas_row:
        secundarias.cell(ofertas_row, 4).value = "Sí"
        secundarias.cell(ofertas_row, 6).value = "compare_at_price > price si está disponible; fallback tag = oferta"
        secundarias.cell(ofertas_row, 10).value = (
            "Crear colección propia para home, menú, SEO y campañas. También puede existir filtro de descuento dentro de categorías, "
            "pero no sustituye la landing de Ofertas."
        )

    marcas = wb["Colecciones marcas"]
    delete_unwanted_rows(marcas, 1, {"Orgie", "Svakom", "Tenga", "ORGIE", "SVAKOM", "TENGA", "Mistress", "MISTRESS"})
    for row in range(4, marcas.max_row + 1):
        brand = marcas.cell(row, 1).value
        if isinstance(brand, str):
            normalized = {"Orgie": "ORGIE", "Svakom": "SVAKOM", "Tenga": "TENGA"}.get(brand, brand.upper())
            marcas.cell(row, 1).value = normalized
            marcas.cell(row, 2).value = normalized.lower()
            marcas.cell(row, 4).value = f"vendor = {normalized}"
    ensure_brand(
        marcas,
        "MISTRESS",
        "mistress",
        "Fetiche y BDSM, accesorios, juegos y experiencias",
        "tipo, material, nivel_iniciacion, color",
        "Crear si la marca/vendor existe en la importación final. Regla automática por vendor exacto.",
    )

    revision = wb["Revisión clienta"]
    next_row = revision.max_row + 1
    copy_row_style(revision, revision.max_row, next_row)
    values = [
        "MVP Shopify",
        "Se crean principales; secundarias solo Novedades y Ofertas; marcas ORGIE, SVAKOM, TENGA y MISTRESS.",
        "Reduce alcance frente a v4 completo.",
        "Documento MVP v1 generado para ejecución inicial en Shopify.",
        "Pendiente validar exportación real de productos y taxonomía limpia de tags/metafields.",
    ]
    for col, value in enumerate(values, 1):
        revision.cell(next_row, col).value = value

    for sheet_name in ("Colecciones principales", "Colecciones secundarias", "Colecciones marcas"):
        ensure_header_fields(wb[sheet_name])

    write_decision_sheet(wb)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 60:
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical or "top",
                        wrap_text=True,
                    )

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
