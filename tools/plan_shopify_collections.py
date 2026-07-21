import csv
import json
from collections import Counter
from pathlib import Path

import openpyxl


BASE = Path(__file__).resolve().parents[1]
WORKBOOK = BASE / "datos" / "07-colecciones" / "LovLory_estructura_colecciones_shopify_MVP_v1.xlsx"
TAXONOMY = BASE / "datos" / "02-datos-intermedios" / "wc-product-export-29-6-2026-marcas-mvp_shopify_taxonomia.csv"
OUTPUT = BASE / "datos" / "06-registros-shopify" / "shopify_collections_creation_plan.json"


def sheet_rows(wb, sheet_name: str) -> list[dict[str, object]]:
    ws = wb[sheet_name]
    headers = [ws.cell(3, col).value for col in range(1, ws.max_column + 1)]
    rows = []
    for row_idx in range(4, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(value not in (None, "") for value in values):
            rows.append(dict(zip(headers, values)))
    return rows


def collection_copy(row: dict[str, object], title_field: str, handle_field: str, rule_field: str) -> dict[str, object]:
    return {
        "title": row.get(title_field),
        "handle": row.get(handle_field),
        "header_eyebrow": row.get("Subtítulo superior cabecera"),
        "header_short_description": row.get("Descripción corta cabecera"),
        "rule": row.get(rule_field),
    }


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    principal = sheet_rows(wb, "Colecciones principales")
    secondary = sheet_rows(wb, "Colecciones secundarias")
    brands = sheet_rows(wb, "Colecciones marcas")

    with TAXONOMY.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    commercial_collections = Counter()
    tags = Counter()
    product_types = Counter()
    vendors = Counter()
    for row in rows:
        vendors[row["Vendor Shopify"]] += 1
        product_types[row["Product type Shopify"]] += 1
        for collection in [item.strip() for item in row["Colecciones comerciales Shopify"].split(",") if item.strip()]:
            commercial_collections[collection] += 1
        for tag in [item.strip() for item in row["Tags Shopify"].split(",") if item.strip()]:
            tags[tag] += 1

    plan = {
        "principal_from_excel": [
            {
                **collection_copy(row, "Colección principal", "Handle sugerido", "Regla de inclusión"),
                "level": row.get("Nivel menú"),
                "block": row.get("Bloque menú"),
            }
            for row in principal
        ],
        "secondary_from_excel": [
            collection_copy(row, "Colección secundaria", "Handle sugerido", "Regla / alimentación")
            for row in secondary
        ],
        "brands_from_excel": [
            collection_copy(row, "Marca / Vendor", "Handle colección marca", "Regla automática")
            for row in brands
        ],
        "product_taxonomy_counts": {
            "commercial_collections": dict(commercial_collections),
            "tags": dict(tags),
            "product_types": dict(product_types),
            "vendors": dict(vendors),
        },
    }
    OUTPUT.write_text(json.dumps(plan, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(plan, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
